import csv
from datetime import datetime, timedelta
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

CSV_PATH = "orders.csv"
OUTPUT_PATH = "Ecommerce_Sales_Dashboard.xlsx"

# ---------- Load & analyze ----------
rows = []
with open(CSV_PATH, newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for r in reader:
        r["order_date_dt"] = datetime.strptime(r["order_date"], "%m/%d/%Y")
        r["acq_dt"] = datetime.strptime(r["customer_acquisition_date"], "%m/%d/%Y")
        r["amount_f"] = float(r["order_amount"])
        rows.append(r)

total_revenue = sum(r["amount_f"] for r in rows)
total_orders = len(rows)
aov = total_revenue / total_orders

monthly = defaultdict(lambda: {"revenue": 0.0, "orders": 0})
for r in rows:
    m = r["order_date_dt"].strftime("%Y-%m")
    monthly[m]["revenue"] += r["amount_f"]
    monthly[m]["orders"] += 1
months_sorted = sorted(monthly.keys())

cat_rev = defaultdict(float)
cat_ord = defaultdict(int)
for r in rows:
    cat_rev[r["product_category"]] += r["amount_f"]
    cat_ord[r["product_category"]] += 1
cats_sorted = sorted(cat_rev, key=lambda c: cat_rev[c], reverse=True)

cust_orders = defaultdict(int)
cust_last = {}
for r in rows:
    cid = r["customer_id"]
    cust_orders[cid] += 1
    if cid not in cust_last or r["order_date_dt"] > cust_last[cid]:
        cust_last[cid] = r["order_date_dt"]

analysis_date = max(r["order_date_dt"] for r in rows)
churn_cutoff = analysis_date - timedelta(days=90)

repeat_customers = [c for c, n in cust_orders.items() if n >= 2]
one_time_customers = [c for c, n in cust_orders.items() if n == 1]
repeat_rate = len(repeat_customers) / len(cust_orders)

churned = [c for c, d in cust_last.items() if d < churn_cutoff]
active = [c for c, d in cust_last.items() if d >= churn_cutoff]
churn_pct = len(churned) / len(cust_orders)

best_month = max(months_sorted, key=lambda m: monthly[m]["revenue"])
worst_month = min(months_sorted, key=lambda m: monthly[m]["revenue"])
top_cat = cats_sorted[0]
bottom_cat = cats_sorted[-1]

def fmt_money(x):
    return f"${x:,.0f}"

def fmt_month(m):
    y, mo = m.split("-")
    names = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    return f"{names[int(mo)-1]} {y}"

# ---------- Styles ----------
NAVY = "1F3864"
BLUE = "2E75B6"
LIGHT_BLUE = "DDEBF7"
GREEN = "548235"
LIGHT_GREEN = "E2EFDA"
PURPLE = "7030A0"
LIGHT_PURPLE = "EDE7F6"
RED = "C00000"
LIGHT_RED = "FBDDDD"
GRAY_FILL = "F2F2F2"

thin = Side(style="thin", color="BFBFBF")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
title_font = Font(name="Calibri", size=20, bold=True, color=NAVY)
sub_font = Font(name="Calibri", size=11, italic=True, color="808080")
label_font_white = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
kpi_font = Font(name="Calibri", size=22, bold=True, color=NAVY)
center = Alignment(horizontal="center", vertical="center")

wb = Workbook()

# ================= Sheet: Raw Data =================
ws_raw = wb.active
ws_raw.title = "Raw Data"
headers = ["customer_id","order_id","order_date","order_amount","product_category","product_name","customer_acquisition_date"]
ws_raw.append(headers)
for r in sorted(rows, key=lambda x: x["order_id"]):
    ws_raw.append([r["customer_id"], r["order_id"], r["order_date"], r["amount_f"],
                   r["product_category"], r["product_name"], r["customer_acquisition_date"]])
ws_raw.freeze_panes = "A2"
ws_raw.auto_filter.ref = f"A1:G{ws_raw.max_row}"
ws_raw.sheet_properties.tabColor = "808080"

# ================= Sheet: Cleaned Data =================
ws_clean = wb.create_sheet("Cleaned Data")
ws_clean.append(headers + ["Order Month"])
for r in sorted(rows, key=lambda x: (x["order_date_dt"], x["order_id"])):
    ws_clean.append([r["customer_id"], r["order_id"], r["order_date_dt"].strftime("%m/%d/%Y"),
                     r["amount_f"], r["product_category"], r["product_name"],
                     r["acq_dt"].strftime("%m/%d/%Y"),
                     r["order_date_dt"].strftime("%Y-%m")])
ws_clean.freeze_panes = "A2"
ws_clean.auto_filter.ref = f"A1:H{ws_clean.max_row}"
ws_clean.sheet_properties.tabColor = BLUE

# ================= Sheet: Sales Analysis =================
ws_sales = wb.create_sheet("Sales Analysis")
ws_sales.sheet_properties.tabColor = GREEN
ws_sales["A1"] = "KEY SALES METRICS"
ws_sales["A1"].font = title_font
kpis_sales = [
    ("Total Revenue", total_revenue, '$#,##0.00'),
    ("Total Orders", total_orders, '#,##0'),
    ("Average Order Value", aov, '$#,##0.00'),
    ("Unique Customers", len(cust_orders), '#,##0'),
    ("Analysis Date (latest order)", analysis_date.strftime("%m/%d/%Y"), None),
]
row = 3
for label, val, numfmt in kpis_sales:
    ws_sales.cell(row=row, column=1, value=label).font = Font(bold=True)
    c = ws_sales.cell(row=row, column=2, value=val)
    if numfmt:
        c.number_format = numfmt
    row += 1
for col, w in zip("AB", (32, 22)):
    ws_sales.column_dimensions[col].width = w

# ================= Sheet: Monthly Analysis =================
ws_m = wb.create_sheet("Monthly Analysis")
ws_m.sheet_properties.tabColor = PURPLE
ws_m.append(["Month", "Revenue", "Orders", "AOV"])
for m in months_sorted:
    rev = monthly[m]["revenue"]; n = monthly[m]["orders"]
    ws_m.append([m, round(rev, 2), n, round(rev / n, 2)])
for cell in ws_m[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=NAVY)
for i in range(2, ws_m.max_row + 1):
    ws_m.cell(row=i, column=2).number_format = '$#,##0.00'
    ws_m.cell(row=i, column=4).number_format = '$#,##0.00'
for col, w in zip("ABCD", (12, 16, 10, 14)):
    ws_m.column_dimensions[col].width = w
ws_m.freeze_panes = "A2"

# ================= Sheet: Category Analysis =================
ws_c = wb.create_sheet("Category Analysis")
ws_c.sheet_properties.tabColor = PURPLE
ws_c.append(["Product Category", "Revenue", "Orders", "% of Revenue"])
for cat in cats_sorted:
    ws_c.append([cat, round(cat_rev[cat], 2), cat_ord[cat], cat_rev[cat] / total_revenue])
for cell in ws_c[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=NAVY)
for i in range(2, ws_c.max_row + 1):
    ws_c.cell(row=i, column=2).number_format = '$#,##0.00'
    ws_c.cell(row=i, column=4).number_format = '0.0%'
for col, w in zip("ABCD", (20, 16, 10, 14)):
    ws_c.column_dimensions[col].width = w
ws_c.freeze_panes = "A2"

# ================= Sheet: Customer Analysis =================
ws_cust = wb.create_sheet("Customer Analysis")
ws_cust.sheet_properties.tabColor = RED
ws_cust["A1"] = "CUSTOMER SUMMARY"
ws_cust["A1"].font = title_font
summary = [
    ("Total Unique Customers", len(cust_orders), '#,##0'),
    ("Repeat Customers (2+ orders)", len(repeat_customers), '#,##0'),
    ("One-Time Customers", len(one_time_customers), '#,##0'),
    ("Repeat Purchase Rate", repeat_rate, '0.0%'),
    ("Active Customers (last 90 days)", len(active), '#,##0'),
    ("Potentially Churned (90+ days)", len(churned), '#,##0'),
    ("Churn Percentage", churn_pct, '0.0%'),
]
r_ = 3
for label, val, nf in summary:
    ws_cust.cell(row=r_, column=1, value=label).font = Font(bold=True)
    cc = ws_cust.cell(row=r_, column=2, value=val)
    if nf: cc.number_format = nf
    r_ += 1

# Status table (feeds pie chart) -> columns D:E
ws_cust["D3"] = "Status"; ws_cust["E3"] = "Customers"
ws_cust["D4"] = "Active";   ws_cust["E4"] = len(active)
ws_cust["D5"] = "Churned";  ws_cust["E5"] = len(churned)
for addr in ("D3", "E3"):
    ws_cust[addr].font = Font(bold=True, color="FFFFFF")
    ws_cust[addr].fill = PatternFill("solid", fgColor=NAVY)

# Customer detail table -> G:J
det_head_row = 3
for j, h in enumerate(["Customer ID", "Orders", "Last Order Date", "Status"], start=7):
    c = ws_cust.cell(row=det_head_row, column=j, value=h)
    c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=NAVY)
rr = det_head_row + 1
for cid in sorted(cust_orders):
    st = "Churned" if cust_last[cid] < churn_cutoff else "Active"
    ws_cust.cell(row=rr, column=7, value=cid)
    ws_cust.cell(row=rr, column=8, value=cust_orders[cid])
    ws_cust.cell(row=rr, column=9, value=cust_last[cid].strftime("%m/%d/%Y"))
    ws_cust.cell(row=rr, column=10, value=st)
    rr += 1
ws_cust.auto_filter.ref = f"G{det_head_row}:J{rr-1}"
for col, w in zip("ABCDEFGHIJ", (34, 14, 4, 4, 12, 4, 14, 10, 16, 12)):
    ws_cust.column_dimensions[col].width = w

# ================= Sheet: Dashboard =================
ws_d = wb.create_sheet("Dashboard", 0)
ws_d.sheet_properties.tabColor = NAVY
ws_d.sheet_view.showGridLines = False
for i in range(1, 26):
    ws_d.column_dimensions[get_column_letter(i)].width = 10.5

ws_d.merge_cells("B2:Y2")
t = ws_d["B2"]; t.value = "E-COMMERCE SALES & CUSTOMER RETENTION DASHBOARD"
t.font = title_font; t.alignment = center
ws_d.row_dimensions[2].height = 30

ws_d.merge_cells("B3:Y3")
s = ws_d["B3"]; s.value = "Data Period: Jul 2024 – Dec 2025  •  6,000 orders  •  2,270 customers"
s.font = sub_font; s.alignment = center

def kpi_card(anchor_col, start_row, label, value, numfmt, color_hex, light_hex, span=3):
    """Draw one KPI card: label row + two merged value rows."""
    c1 = get_column_letter(anchor_col); c2 = get_column_letter(anchor_col + span - 1)
    ws_d.merge_cells(f"{c1}{start_row}:{c2}{start_row}")
    lc = ws_d[f"{c1}{start_row}"]
    lc.value = label
    lc.font = label_font_white
    lc.fill = PatternFill("solid", fgColor=color_hex)
    lc.alignment = center
    ws_d.merge_cells(f"{c1}{start_row+1}:{c2}{start_row+2}")
    vc = ws_d[f"{c1}{start_row+1}"]
    vc.value = value
    vc.font = Font(name="Calibri", size=20, bold=True, color=color_hex)
    vc.fill = PatternFill("solid", fgColor=light_hex)
    vc.alignment = center
    if numfmt:
        vc.number_format = numfmt
    # borders around the card
    for rr_ in range(start_row, start_row + 3):
        for cc_ in range(anchor_col, anchor_col + span):
            ws_d.cell(row=rr_, column=cc_).border = border_all
    ws_d.row_dimensions[start_row].height = 18
    ws_d.row_dimensions[start_row + 1].height = 22
    ws_d.row_dimensions[start_row + 2].height = 22

# Row block 1: main KPIs
kpi_card(2, 5, "TOTAL REVENUE", round(total_revenue, 2), '$#,##0', NAVY, LIGHT_BLUE)
kpi_card(6, 5, "TOTAL ORDERS", total_orders, '#,##0', BLUE, LIGHT_BLUE)
kpi_card(10, 5, "AVG ORDER VALUE", round(aov, 2), '$#,##0.00', GREEN, LIGHT_GREEN)
# Row block 2: retention KPIs
kpi_card(2, 9, "REPEAT PURCHASE RATE", repeat_rate, '0.0%', PURPLE, LIGHT_PURPLE)
kpi_card(6, 9, "CHURNED CUSTOMERS (90d+)", len(churned), '#,##0', RED, LIGHT_RED, span=3)
kpi_card(10, 9, "UNIQUE CUSTOMERS", len(cust_orders), '#,##0', NAVY, LIGHT_BLUE)

# ---- Chart 1: Monthly revenue line ----
lc1 = LineChart()
lc1.title = "Monthly Revenue Trend"
lc1.style = 12
lc1.y_axis.title = "Revenue ($)"
lc1.x_axis.title = "Month"
data = Reference(ws_m, min_col=2, min_row=1, max_row=ws_m.max_row)
catsref = Reference(ws_m, min_col=1, min_row=2, max_row=ws_m.max_row)
lc1.add_data(data, titles_from_data=True)
lc1.set_categories(catsref)
lc1.series[0].smooth = False
lc1.width = 17.5
lc1.height = 9
ws_d.add_chart(lc1, "B13")

# ---- Chart 2: Orders per month column ----
bc1 = BarChart()
bc1.type = "col"
bc1.title = "Orders per Month"
bc1.style = 10
bc1.y_axis.title = "Orders"
data = Reference(ws_m, min_col=3, min_row=1, max_row=ws_m.max_row)
bc1.add_data(data, titles_from_data=True)
bc1.set_categories(catsref)
bc1.legend = None
bc1.width = 17.5
bc1.height = 9
ws_d.add_chart(bc1, "L13")

# ---- Chart 3: Revenue by category bar ----
bc2 = BarChart()
bc2.type = "bar"
bc2.title = "Revenue by Product Category"
bc2.style = 11
data = Reference(ws_c, min_col=2, min_row=1, max_row=ws_c.max_row)
catsref2 = Reference(ws_c, min_col=1, min_row=2, max_row=ws_c.max_row)
bc2.add_data(data, titles_from_data=True)
bc2.set_categories(catsref2)
bc2.legend = None
bc2.width = 17.5
bc2.height = 9
ws_d.add_chart(bc2, "B33")

# ---- Chart 4: Customer status pie ----
pc = PieChart()
pc.title = "Customer Status (Active vs Churned)"
data = Reference(ws_cust, min_col=5, min_row=3, max_row=5)
labels = Reference(ws_cust, min_col=4, min_row=4, max_row=5)
pc.add_data(data, titles_from_data=True)
pc.set_categories(labels)
pc.dataLabels = DataLabelList()
pc.dataLabels.showPercent = True
pc.width = 17.5
pc.height = 9
ws_d.add_chart(pc, "L33")

# ---- Key insights box ----
insight_row = 53
ws_d.merge_cells(f"B{insight_row}:Y{insight_row}")
h = ws_d[f"B{insight_row}"]
h.value = "KEY INSIGHTS"
h.font = Font(size=13, bold=True, color="FFFFFF")
h.fill = PatternFill("solid", fgColor=NAVY)
h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_d.row_dimensions[insight_row].height = 20

insights = (
    f"1. Best month: {fmt_month(best_month)} ({fmt_money(monthly[best_month]['revenue'])}); "
    f"weakest month: {fmt_month(worst_month)} ({fmt_money(monthly[worst_month]['revenue'])}).\n"
    f"2. {top_cat} leads categories with {fmt_money(cat_rev[top_cat])} "
    f"({cat_rev[top_cat]/total_revenue:.0%} of revenue); {bottom_cat} is lowest at {fmt_money(cat_rev[bottom_cat])}.\n"
    f"3. {len(repeat_customers):,} of {len(cust_orders):,} customers ordered more than once "
    f"(repeat purchase rate {repeat_rate:.1%}).\n"
    f"4. {len(churned):,} customers ({churn_pct:.1%}) haven't ordered in 90+ days — target them "
    f"with a win-back campaign.\n"
    f"5. Average order value is {fmt_money(aov)} across {total_orders:,} orders."
)
ws_d.merge_cells(f"B{insight_row+1}:Y{insight_row+6}")
ib = ws_d[f"B{insight_row+1}"]
ib.value = insights
ib.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
ib.font = Font(size=11)
ib.fill = PatternFill("solid", fgColor=GRAY_FILL)
for rr_ in range(insight_row, insight_row + 7):
    for cc_ in range(2, 26):
        ws_d.cell(row=rr_, column=cc_).border = border_all

wb.save(OUTPUT_PATH)
print(f"Saved {OUTPUT_PATH}")

print("\n===== SUMMARY =====")
print(f"Total Revenue:      {fmt_money(total_revenue)}")
print(f"Total Orders:       {total_orders:,}")
print(f"Average Order Value:{fmt_money(aov)}")
print(f"Unique Customers:   {len(cust_orders):,}")
print(f"Repeat Rate:        {repeat_rate:.1%} ({len(repeat_customers):,} customers)")
print(f"Churned (90d+):     {len(churned):,} ({churn_pct:.1%})")
print(f"Best Month:         {best_month} ({fmt_money(monthly[best_month]['revenue'])})")
print(f"Worst Month:        {worst_month} ({fmt_money(monthly[worst_month]['revenue'])})")
print(f"Top Category:       {top_cat} ({fmt_money(cat_rev[top_cat])})")
print(f"Bottom Category:    {bottom_cat} ({fmt_money(cat_rev[bottom_cat])})")
